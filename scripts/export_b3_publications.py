#!/usr/bin/env python3
"""
Export Panel B3 publications to an Excel workbook, one sheet per country.

Columns per row (one publication per row):
    1. In-text citation (mirrors b3InTextCitation() in js/itaukei-database.js)
    2. Full reference (APA-flavoured, assembled from Zotero fields)
    3. iTaukei lead author (empty when the lead is not iTaukei)
    4. iTaukei co-authors (semicolon-separated; empty when none)

Rows are shaded with alternating fill for readability.

Data sources (all under vave-lab/data/):
    - itaukei-zotero-snapshot.json  (items + collections; "Where study was done"
      collection tree drives the country mapping)
    - scholar-profiles.json         (canonical iTaukei scholars + name aliases +
      hard "notItaukeiAuthors" list)
    - progress-roster (fetched live from the Apps Script endpoint used by the
      dashboard so the export uses the same canonical set as the map)

Usage:
    cd vave-lab
    VAVELAB_PASSCODE='Arachnid1!' python3 scripts/decrypt_data.py --all
    python3 scripts/export_b3_publications.py \
        --out /home/user/workspace/b3_publications_by_country.xlsx
    rm -f data/*.json  # remember to re-clean plaintext when done
"""

from __future__ import annotations

import argparse
import json
import re
import sys
import urllib.request
from pathlib import Path
from typing import Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


REPO = Path(__file__).resolve().parent.parent
DATA = REPO / "data"

PROGRESS_ROSTER_URL = (
    "https://script.google.com/macros/s/"
    "AKfycbxNBAxheCV29gxktJjvC3xNYhnUDN4JDk-nUdrF3ckdkdgZ6NXoD432avysXY64itAf/"
    "exec?mode=progress"
)

# Same set as B3_COUNTRY_COORDS in js/itaukei-database.js. Only used for the
# "Fiji Provinces" → "Fiji (Provinces)" display transform; the export uses
# Zotero collection names directly for the sheet names/order otherwise.
DISPLAY_NAME_OVERRIDES = {
    "Fiji Provinces": "Fiji (Provinces)",
}

# Column widths (chars). Tuned so citation + names fit without wrapping and
# the full reference wraps within a comfortable width.
COL_WIDTHS = [26, 90, 30, 45]

# Alternating row fills — two soft, high-contrast tints.
FILL_A = PatternFill("solid", fgColor="FFFFFFFF")  # white
FILL_B = PatternFill("solid", fgColor="FFF3F0EA")  # warm off-white
HEADER_FILL = PatternFill("solid", fgColor="FF712B13")  # rust
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFFFF")
BODY_FONT = Font(name="Calibri", size=11)
LEAD_FONT = Font(name="Calibri", size=11, bold=True, color="FF712B13")  # rust
THIN = Side(border_style="thin", color="FFDDDDDD")
CELL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


# ---------- Name / citation helpers (mirrors the JS logic) ----------

_HYPH_RE = re.compile(r"[\u2010\u2011\u2013\u2212]")


def keyify(name: str) -> str:
    if not name:
        return ""
    s = _HYPH_RE.sub("-", name).lower()
    return re.sub(r"[^a-z0-9]", "", s)


def surname_of(creator: str) -> str:
    """Extract the surname from a creator string.

    Handles both "Last, First" and "First Last" forms.
    """
    if "," in creator:
        return creator.split(",", 1)[0].strip()
    tokens = creator.strip().split()
    return tokens[-1] if tokens else creator


def normalize_lastfirst(creator: str) -> str:
    """Return the creator in canonical "Last, First" form.

    Mirrors what the JS does when checking canonical keys.
    """
    s = creator.strip()
    if "," in s:
        return s
    tokens = s.split()
    if len(tokens) >= 2:
        return tokens[-1] + ", " + " ".join(tokens[:-1])
    return s


def display_firstlast(creator: str) -> str:
    """Return "First Last" for display in the spreadsheet columns."""
    s = creator.strip()
    if "," in s:
        parts = s.split(",", 1)
        last = parts[0].strip()
        first = (parts[1].strip() if len(parts) > 1 else "")
        return f"{first} {last}".strip() if first else last
    return s


# ---------- iTaukei canonical set ----------


def build_itaukei_keys(profiles: dict, roster_rows: List[dict]) -> set:
    """Union of:
    - every canonical scholar name in scholar-profiles.json
    - every alias key AND alias value in nameAliases
    - every canonical / built / flipped variant from the progress-Sheet roster

    Mirrors the JS union used by creatorIsItaukei() at runtime.
    """
    canon: set = set()

    for s in profiles.get("scholars", []):
        name = s.get("name") or ""
        if name:
            canon.add(keyify(name))

    for variant, canonical in (profiles.get("nameAliases") or {}).items():
        canon.add(keyify(variant))
        canon.add(keyify(canonical))

    for row in roster_rows or []:
        if not row:
            continue
        canonical = (
            row.get("canonical") or row.get("canonicalName")
            or row.get("name") or row.get("scholar") or ""
        )
        first = row.get("firstName") or row.get("first") or ""
        last = row.get("lastName") or row.get("last") or ""
        built = f"{last}, {first}" if last and first else (last or first or "")
        flipped = f"{first} {last}" if first and last else ""
        for candidate in (canonical, built, flipped):
            k = keyify(candidate)
            if k:
                canon.add(k)

    canon.discard("")
    return canon


def creator_is_itaukei(name: str, canon: set) -> bool:
    """Match the multi-form fallback logic of creatorIsItaukei() in JS."""
    if not name or not canon:
        return False
    if keyify(name) in canon:
        return True
    s = name.strip()
    if "," not in s:
        tokens = s.split()
        if len(tokens) >= 2:
            flipped = tokens[-1] + ", " + " ".join(tokens[:-1])
            if keyify(flipped) in canon:
                return True
    else:
        parts = s.split(",", 1)
        first_tok = (parts[1].strip().split() or [""])[0]
        if first_tok:
            trimmed = parts[0].strip() + ", " + first_tok
            if keyify(trimmed) in canon:
                return True
    return False


# ---------- Citation formatting ----------


def in_text_citation(item: dict) -> str:
    creators = item.get("creators") or []
    year = item.get("year") or "n.d."
    if not creators:
        return f"Unknown ({year})"
    first = creators[0]
    surname = surname_of(first)
    if len(creators) == 1:
        return f"{surname} ({year})"
    if len(creators) == 2:
        return f"{surname} & {surname_of(creators[1])} ({year})"
    return f"{surname} et al. ({year})"


def author_list_apa(creators: List[str]) -> str:
    """Return an author list in "Last, F. M., Last, F. M., & Last, F. M." style.

    Falls back to whatever form Zotero gave us if we can't split a first name.
    """
    if not creators:
        return ""
    parts: List[str] = []
    for c in creators:
        s = c.strip()
        if "," in s:
            last, first = [t.strip() for t in s.split(",", 1)]
            initials = " ".join(
                (t[0] + ".") for t in re.split(r"\s+|-", first) if t
            )
            parts.append(f"{last}, {initials}".strip() if initials else last)
        else:
            tokens = s.split()
            if len(tokens) >= 2:
                last = tokens[-1]
                initials = " ".join((t[0] + ".") for t in tokens[:-1] if t)
                parts.append(f"{last}, {initials}".strip())
            else:
                parts.append(s)
    if len(parts) == 1:
        return parts[0]
    if len(parts) == 2:
        return f"{parts[0]}, & {parts[1]}"
    return ", ".join(parts[:-1]) + ", & " + parts[-1]


def full_reference(item: dict) -> str:
    """Assemble an APA-ish reference from Zotero fields.

    Deliberately conservative — we only surface fields present on the item.
    Not a full CSL citation processor.
    """
    authors = author_list_apa(item.get("creators") or [])
    year = item.get("year") or item.get("date") or "n.d."
    title = (item.get("title") or "").strip()
    item_type = item.get("itemType") or ""

    parts: List[str] = []
    if authors:
        parts.append(f"{authors} ({year}).")
    else:
        parts.append(f"({year}).")

    if title:
        # Titles for articles are un-italicised; theses/books italicised — we
        # can't render italics in a plaintext cell so we just include the title.
        parts.append(f"{title}.")

    if item_type == "journalArticle":
        pub = (item.get("publicationTitle") or "").strip()
        if pub:
            parts.append(f"{pub}.")
    elif item_type == "thesis":
        thesis_type = (item.get("thesisType") or "Thesis").strip()
        university = (item.get("university") or "").strip()
        tail = thesis_type + (f", {university}" if university else "")
        parts.append(tail + ".")
    elif item_type == "bookSection":
        pub = (item.get("publicationTitle") or item.get("bookTitle") or "").strip()
        if pub:
            parts.append(f"In {pub}.")
    elif item_type == "book":
        publisher = (item.get("publisher") or "").strip()
        if publisher:
            parts.append(f"{publisher}.")
    elif item_type == "report":
        publisher = (item.get("publisher") or item.get("institution") or "").strip()
        if publisher:
            parts.append(f"{publisher}.")

    doi = (item.get("DOI") or "").strip()
    url = (item.get("url") or "").strip()
    if doi:
        # Zotero sometimes stores the DOI already URL-prefixed; don't double it.
        if doi.startswith("http://") or doi.startswith("https://"):
            parts.append(doi)
        else:
            parts.append(f"https://doi.org/{doi}")
    elif url:
        parts.append(url)

    return " ".join(p for p in parts if p).strip()


# ---------- Country bucketing (mirrors initB3Map) ----------


def build_country_buckets(snap: dict, canon: set) -> Dict[str, List[dict]]:
    """Return { country_name: [item, ...] } including all items whose
    Zotero collection tree resolves under a "Where study was done" child.

    Each item is annotated in-place with ``__leadIsItaukei`` (bool) and
    ``__coauthItaukei`` (list of iTaukei co-author display names) so the
    caller doesn't re-run the classification.
    """
    cols = snap.get("collections") or []
    by_key = {c["key"]: c for c in cols}
    where_root = next(
        (c for c in cols if c.get("name") == "Where study was done" and not c.get("parent")),
        by_key.get("V3HLPDPL"),
    )
    if not where_root:
        raise RuntimeError('Zotero "Where study was done" collection not found.')

    country_of_key: Dict[str, str] = {}
    countries: List[str] = []
    for country in [c for c in cols if c.get("parent") == where_root["key"]]:
        countries.append(country["name"])
        country_of_key[country["key"]] = country["name"]
        stack = [c for c in cols if c.get("parent") == country["key"]]
        while stack:
            c = stack.pop(0)
            country_of_key[c["key"]] = country["name"]
            stack.extend([x for x in cols if x.get("parent") == c["key"]])

    buckets: Dict[str, List[dict]] = {name: [] for name in countries}
    for item in snap.get("items") or []:
        hits: set = set()
        for k in item.get("collections") or []:
            cn = country_of_key.get(k)
            if cn:
                hits.add(cn)
        if not hits:
            continue
        creators = item.get("creators") or []
        lead_itaukei = bool(creators) and creator_is_itaukei(creators[0], canon)
        # Order-preserving de-dup — Zotero occasionally lists the same author
        # twice on an item (e.g. "R. Varea; R. Varea"); we don't want that
        # duplication to leak into the export column.
        coauths: List[str] = []
        seen_keys: set = set()
        for c in creators[1:]:
            if not creator_is_itaukei(c, canon):
                continue
            k = keyify(c)
            if k in seen_keys:
                continue
            seen_keys.add(k)
            coauths.append(display_firstlast(c))
        item["__leadIsItaukei"] = lead_itaukei
        item["__coauthItaukei"] = coauths
        for cn in hits:
            buckets[cn].append(item)

    return buckets


# ---------- Excel writing ----------


def safe_sheet_name(name: str) -> str:
    """Excel sheet name limits: 31 chars, no [ ] : * ? / \\ or leading/trailing '."""
    bad = re.compile(r"[\[\]:*?/\\]")
    cleaned = bad.sub(" ", name).strip().strip("'")
    return cleaned[:31] or "Sheet"


def write_workbook(buckets: Dict[str, List[dict]], out_path: Path) -> None:
    wb = Workbook()
    # Remove the default sheet — we'll create a Summary as the first sheet.
    default = wb.active
    wb.remove(default)

    summary = wb.create_sheet("Summary")
    summary.append(["Country", "Publications", "iTaukei-led", "Others-led"])
    for cell in summary[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = CELL_BORDER

    # Deterministic country order: descending by count, alphabetical tiebreak.
    country_order = sorted(
        buckets.keys(),
        key=lambda cn: (-len(buckets[cn]), cn.lower()),
    )

    for country in country_order:
        items = buckets[country]
        if not items:
            continue

        display = DISPLAY_NAME_OVERRIDES.get(country, country)
        led = sum(1 for it in items if it.get("__leadIsItaukei"))
        others = len(items) - led

        summary.append([display, len(items), led, others])

        ws = wb.create_sheet(safe_sheet_name(display))
        headers = [
            "In-text citation",
            "Full reference",
            "iTaukei lead author",
            "iTaukei co-authors",
        ]
        ws.append(headers)
        for i, cell in enumerate(ws[1], start=1):
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True,
            )
            cell.border = CELL_BORDER
            ws.column_dimensions[get_column_letter(i)].width = COL_WIDTHS[i - 1]

        ws.row_dimensions[1].height = 28
        ws.freeze_panes = "A2"

        # Sort newest first inside each country, same as the map popup.
        sorted_items = sorted(items, key=lambda it: (-(it.get("year") or 0),))

        for idx, item in enumerate(sorted_items, start=1):
            cite = in_text_citation(item)
            ref = full_reference(item)
            creators = item.get("creators") or []
            lead_author = (
                display_firstlast(creators[0])
                if creators and item.get("__leadIsItaukei") else ""
            )
            coauth = "; ".join(item.get("__coauthItaukei") or [])
            row = [cite, ref, lead_author, coauth]
            ws.append(row)

            excel_row = idx + 1  # +1 for header
            fill = FILL_A if (idx % 2 == 1) else FILL_B
            for col_idx, _ in enumerate(row, start=1):
                cell = ws.cell(row=excel_row, column=col_idx)
                cell.fill = fill
                cell.border = CELL_BORDER
                cell.alignment = Alignment(
                    vertical="top",
                    wrap_text=(col_idx == 2),  # only full reference wraps
                    horizontal="left",
                )
                cell.font = LEAD_FONT if col_idx == 3 and lead_author else BODY_FONT

        # Sheet-wide niceties.
        ws.sheet_view.showGridLines = False

    # Style Summary rows with alternating fill too.
    for i in range(2, summary.max_row + 1):
        fill = FILL_A if (i % 2 == 0) else FILL_B
        for j in range(1, 5):
            c = summary.cell(row=i, column=j)
            c.fill = fill
            c.border = CELL_BORDER
            c.font = BODY_FONT
            c.alignment = Alignment(
                horizontal="left" if j == 1 else "right", vertical="center",
            )
    for j, w in enumerate([28, 16, 16, 16], start=1):
        summary.column_dimensions[get_column_letter(j)].width = w
    summary.freeze_panes = "A2"
    summary.sheet_view.showGridLines = False

    wb.save(out_path)


# ---------- Entrypoint ----------


def load_progress_roster() -> List[dict]:
    """Fetch the same live roster the dashboard uses.

    Falls back to an empty list on network failure (the profiles-only union
    still produces good coverage — just slightly narrower).
    """
    try:
        with urllib.request.urlopen(PROGRESS_ROSTER_URL, timeout=15) as r:
            data = json.load(r)
    except Exception as exc:  # noqa: BLE001
        print(f"WARN: could not fetch progress roster: {exc}", file=sys.stderr)
        return []
    if isinstance(data, list):
        return data
    if isinstance(data, dict):
        for key in ("rows", "result"):
            if isinstance(data.get(key), list):
                return data[key]
    return []


def main() -> int:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--out", default=str(REPO / "b3_publications_by_country.xlsx"))
    args = p.parse_args()

    snap_path = DATA / "itaukei-zotero-snapshot.json"
    prof_path = DATA / "scholar-profiles.json"
    if not snap_path.exists() or not prof_path.exists():
        print(
            "ERROR: data/*.json not found. Decrypt first with\n"
            "  VAVELAB_PASSCODE='...' python3 scripts/decrypt_data.py --all",
            file=sys.stderr,
        )
        return 2

    snap = json.load(snap_path.open())
    profiles = json.load(prof_path.open())
    roster = load_progress_roster()

    canon = build_itaukei_keys(profiles, roster)
    print(f"iTaukei canonical keys: {len(canon)}")

    buckets = build_country_buckets(snap, canon)
    total_items = sum(len(v) for v in buckets.values())
    total_led = sum(
        1 for items in buckets.values() for it in items if it.get("__leadIsItaukei")
    )
    print(
        f"countries with publications: {sum(1 for v in buckets.values() if v)}; "
        f"total publications: {total_items}; iTaukei-led: {total_led}; "
        f"others-led: {total_items - total_led}"
    )

    out_path = Path(args.out)
    write_workbook(buckets, out_path)
    print(f"WROTE {out_path}  ({out_path.stat().st_size:,} bytes)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
